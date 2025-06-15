import os
import polars as pl
from openpyxl import load_workbook
import json
import pandas as pd
from utils import load_api_data, create_client, get_response

def list_report_files(folder_path):
    """
    指定フォルダ以下の「報告書」を含む.xlsxファイルのパスをリストアップする。
    """
    report_files = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith('.xlsx') and '報告書' in file:
                full_path = os.path.join(root, file)
                report_files.append(full_path)
    return report_files

def read_report_to_df(file_path):
    """
    Excelファイルをopenpyxlで読み込み、C7, C9, C11セルの値（日付、氏名、状況）を抽出してpolars DataFrameに変換する。
    """
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active

    # セルから値を取得
    date_value = ws['C7'].value
    name_value = ws['C9'].value
    status_value = ws['C11'].value

    # polars DataFrameに1行でまとめる
    pl_df = pl.DataFrame(
        [[date_value, name_value, status_value]],
        schema=['日付', '氏名', '状況']
    )
    return pl_df

def extract_delays_statuses(status_list, client, model):
    """
    状況リストから工程遅延しそうな内容をAzureOpenAIで抽出する関数
    """
    # 状況を改行区切りでまとめる
    content = "以下の状況説明の中で工程遅延しそうな内容だけを抽出してください。\n\n"
    for i, status in enumerate(status_list, 1):
        content += f"{i}. {status}\n"

    # AzureOpenAIに問い合わせ
    response = get_response(client, model, content)
    return response

def extract_delay_risk_for_status(status, client, model):
    """
    1つの状況説明に対して工程遅延リスクを判定し、リスク内容を返す関数。
    リスクがなければ「なし」を返す。
    """
    prompt = f"以下の状況説明について、工程遅延リスクがある場合はその内容を簡潔に記述し、なければ「なし」とだけ答えてください。\n\n状況説明: {status}"
    response = get_response(client, model, prompt)
    # 応答の前後の空白を除去
    return response.strip()

def main(folder=None, api_config_path="api_gpt4o.json", output_path="工程遅延リスク報告.xlsx"):
    if folder is None:
        folder = os.getcwd()

    # API設定読み込みとクライアント作成
    api_data = load_api_data(api_config_path)
    client, model = create_client(api_data)

    files = list_report_files(folder)
    dfs = []
    for f in files:
        df = read_report_to_df(f)
        dfs.append(df)
    if dfs:
        combined_df = pl.concat(dfs)
        print("=== 全状況データ ===")
        print(combined_df)

        # polars DataFrameをpandas DataFrameに変換
        pdf = combined_df.to_pandas()

        # 各行の"状況"に対して遅延リスク判定を実施
        delay_risks = []
        for status in pdf["状況"]:
            risk = extract_delay_risk_for_status(status, client, model)
            delay_risks.append(risk)

        # 新しいカラム"工程遅延リスク"を追加
        pdf["工程遅延リスク"] = delay_risks

        # Excelファイルに保存
        pdf.to_excel(output_path, index=False)

        print(f"\n=== 工程遅延リスクを含む報告をExcelファイルに保存しました: {output_path} ===")
    else:
        print("報告書ファイルが見つかりませんでした。")

if __name__ == "__main__":
    main()
