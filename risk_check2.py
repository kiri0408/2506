import os
from openpyxl import load_workbook, Workbook
from utils import load_api_data, create_client, get_response

# 作業フォルダと対象フォルダのパス
WORK_DIR = os.getcwd()
TARGET_DIR = os.path.join(WORK_DIR, "output_reports")
OUTPUT_FILE = os.path.join(WORK_DIR, "risk_check_output.xlsx")

# API接続情報ファイルパス（適宜変更してください）
API_CONFIG_PATH = os.path.join(WORK_DIR, "api_gpt4o.json")

def list_target_files(folder):
    """指定フォルダ内のファイル名に「報告書」を含む.xlsxファイルをリストアップ"""
    files = []
    for root, dirs, filenames in os.walk(folder):
        for filename in filenames:
            if "報告書" in filename and filename.endswith(".xlsx") and not filename.startswith("~$"):
                files.append(os.path.join(root, filename))
    return files

def check_risk(client, model, situation_text):
    """状況テキストをAzureOpenAIに投げて工程遅延リスクを判定"""
    prompt = (
        "以下の状況説明を読み、工程遅延リスクがある場合は「工程遅延リスク」として"
        "その内容を簡潔に記述してください。リスクがなければ「なし」とだけ答えてください。\n\n"
        f"状況説明:\n{situation_text}"
    )
    response = get_response(client, model, prompt)
    return response.strip()

def main():
    # API情報読み込みとクライアント作成
    api_data = load_api_data(API_CONFIG_PATH)
    client, model = create_client(api_data)

    # 対象ファイル取得
    target_files = list_target_files(TARGET_DIR)

    # 結果格納用リスト
    results = []

    for file_path in target_files:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # セルから値取得
        date_val = ws["C7"].value
        name_val = ws["C9"].value
        situation_val = ws["C11"].value

        # 工程遅延リスク判定
        risk_val = check_risk(client, model, situation_val)

        results.append([date_val, name_val, situation_val, risk_val])

    # 結果をExcelに出力
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "工程遅延リスクチェック結果"

    # ヘッダー
    out_ws.append(["日付", "氏名", "状況", "工程遅延リスク"])

    # データ書き込み
    for row in results:
        out_ws.append(row)

    # 保存
    out_wb.save(OUTPUT_FILE)
    print(f"結果を {OUTPUT_FILE} に保存しました。")

if __name__ == "__main__":
    main()
